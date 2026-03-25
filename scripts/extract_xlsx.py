"""
scripts/extract_xlsx.py
-----------------------
Run once to extract all embedded images from the Excel file
and build the initial dataset JSON.

Usage:
    python scripts/extract_xlsx.py \
        --xlsx  data/Export_-_foto_s_HWL_Without_Lebel.xlsx \
        --out   data/hwl_images \
        --json  data/hwl_dataset.json
"""

import argparse
import json
import os
import shutil
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import load_workbook


def extract(xlsx_path: str, images_dir: str, dataset_json: str) -> None:
    print(f"[extract] Opening {xlsx_path} ...")
    os.makedirs(images_dir, exist_ok=True)

    with zipfile.ZipFile(xlsx_path, "r") as z:
        # ── 1. rId → image filename ─────────────────────────────────
        rels: dict[str, str] = {}
        rel_xml = "xl/drawings/_rels/drawing1.xml.rels"
        tree = ET.parse(z.open(rel_xml))
        for elem in tree.iter():
            if elem.tag.endswith("Relationship"):
                rid    = elem.attrib.get("Id")
                target = elem.attrib.get("Target", "")
                if "media" in target:
                    rels[rid] = target.split("/")[-1]   # e.g. "image1.png"

        # ── 2. Excel row → image filename ───────────────────────────
        row_to_img: dict[int, str] = {}
        ns = {
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a":   "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r":   "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        tree = ET.parse(z.open("xl/drawings/drawing1.xml"))
        r_embed = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"

        for anchor in (
            tree.findall(".//xdr:twoCellAnchor", ns) +
            tree.findall(".//xdr:oneCellAnchor", ns)
        ):
            from_elem = anchor.find("xdr:from", ns)
            row = None
            if from_elem is not None:
                row_elem = from_elem.find("xdr:row", ns)
                if row_elem is not None:
                    row = int(row_elem.text) + 1   # 0-indexed → 1-indexed

            blip = anchor.find(".//a:blip", ns)
            if blip is not None and row is not None:
                rid = blip.attrib.get(r_embed)
                if rid:
                    row_to_img[row] = rels.get(rid, "")

        # ── 3. Read material rows from Sheet1 ──────────────────────
        wb = load_workbook(xlsx_path, read_only=True)
        ws = wb["Sheet1"]
        rows = list(ws.iter_rows(values_only=True))

        dataset: list[dict] = []
        skipped = 0

        for row_idx, row in enumerate(rows[1:], start=2):   # skip header row 1
            material_id = row[2]
            img_filename = row_to_img.get(row_idx)

            if not material_id or not img_filename:
                skipped += 1
                continue

            mat_dir = os.path.join(images_dir, str(material_id))
            os.makedirs(mat_dir, exist_ok=True)
            dst = os.path.join(mat_dir, "image_1.png")

            with z.open(f"xl/media/{img_filename}") as src_f:
                with open(dst, "wb") as dst_f:
                    shutil.copyfileobj(src_f, dst_f)

            dataset.append({
                "material_id":      str(material_id),
                "storage_type":     row[0],
                "storage_bin":      row[1],
                "total_stock":      row[3],
                "storage_location": row[4],
                "plant":            row[6],
                "duration":         row[7],
                "image_path":       dst,
            })

    with open(dataset_json, "w") as f:
        json.dump(dataset, f, indent=2)

    print(f"[extract] ✅  {len(dataset)} materials extracted  (skipped {skipped})")
    print(f"           Images  → {images_dir}/")
    print(f"           Dataset → {dataset_json}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract images from HWL xlsx")
    parser.add_argument("--xlsx", default="data/Export - foto's HWL Without Lebel.xlsx")
    parser.add_argument("--out",  default="data/hwl_images")
    parser.add_argument("--json", default="data/hwl_dataset.json")
    args = parser.parse_args()
    extract(args.xlsx, args.out, args.json)


if __name__ == "__main__":
    main()